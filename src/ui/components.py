"""
Streamlit UI components
"""

import streamlit as st
import pandas as pd
import logging
from pathlib import Path
from calendar import month_name

from .formatters import parse_split_csv, calculate_summary_stats, transform_to_multiindex

logger = logging.getLogger(__name__)


def show_config_error(error_msg: str):
    """Display configuration error with helpful instructions"""
    st.error(":warning: Missing Configuration")
    st.markdown("""
    Please set these environment variables:
    ```
    JIRA_URL=https://your-company.atlassian.net
    JIRA_USERNAME=your-email@company.com
    JIRA_API_TOKEN=your-api-token
    ```
    
    **For Docker:** Make sure your `.env` file exists and restart:
    ```bash
    make web-stop
    make web
    ```
    """)
    st.error(f"Error details: {error_msg}")
    st.stop()


def _display_dataframe_with_styling(df: pd.DataFrame, project_col, is_multilevel: bool = False):
    """Helper to display dataframe with TOTAL row styling"""
    
    def highlight_total(row):
        try:
            return ['background-color: #f0f2f6; font-weight: bold' if row.get(project_col) == 'TOTAL' else '' for _ in row]
        except:
            return ['' for _ in row]
    
    def format_number(val):
        if val is None or val == '' or val == '-':
            return '-'
        try:
            return f'{float(val):.1f}'
        except (ValueError, TypeError):
            return str(val)
    
    # Build format dict based on column type
    if is_multilevel:
        format_dict = {col: format_number for col in df.columns 
                      if not str(col[0]).startswith('Project') and not str(col[1]).startswith('Component')}
    else:
        format_dict = {col: format_number for col in df.columns 
                      if not str(col).startswith('Project') and not str(col).startswith('Component')}
    
    
    # Create column config to pin Project and Component columns
    column_config = {}
    if is_multilevel:
        # For MultiIndex columns, use integer indices
        # Note: Must account for DataFrame index levels even with hide_index=True
        # Index 0 is the hidden index, so Project=1, Component=2
        num_index_levels = df.index.nlevels
        column_config[num_index_levels] = st.column_config.Column(pinned=True)  # Project
        column_config[num_index_levels + 1] = st.column_config.Column(pinned=True)  # Component
    else:
        # Find the actual column names for Project and Component
        project_cols = [col for col in df.columns if str(col).startswith('Project')]
        component_cols = [col for col in df.columns if str(col).startswith('Component')]
        if project_cols:
            column_config[project_cols[0]] = st.column_config.Column(pinned=True)
        if component_cols:
            column_config[component_cols[0]] = st.column_config.Column(pinned=True)
    
    try:
        styled_df = df.style.apply(highlight_total, axis=1).format(format_dict)
        st.dataframe(styled_df, use_container_width=True, hide_index=True, height=250, column_config=column_config if column_config else None)
    except Exception:
        # Fallback: display without styling
        st.dataframe(df, use_container_width=True, hide_index=True, height=250, column_config=column_config if column_config else None)


def _parse_xlsx_sheet(ws, is_multilevel: bool = False):
    """Parse XLSX sheet into dev and maint dataframes, and extract metadata
    
    Returns:
        tuple: (dev_data, maint_data, header_row, header_row_2, metadata)
    """
    dev_data = []
    maint_data = []
    current_section = None
    header_row = None
    header_row_2 = None
    metadata = {}
    
    for row in ws.iter_rows(values_only=True):
        # Skip empty rows
        if not any(row):
            continue
        
        # Parse metadata lines (before sections start)
        if row[0] and isinstance(row[0], str):
            if row[0].startswith('Generated:'):
                metadata['generated'] = row[0].replace('Generated:', '').strip()
                continue
        
        # Check for section headers
        if row[0] == 'DEVELOPMENT':
            current_section = 'dev'
            header_row = None
            header_row_2 = None
            continue
        elif row[0] == 'MAINTENANCE':
            current_section = 'maint'
            header_row = None
            header_row_2 = None
            continue
        
        # Capture header rows
        if current_section and header_row is None and row[0] in ['Project', 'TOTAL']:
            if row[0] == 'Project':
                header_row = row
            continue
        
        # Capture second header row for weekly
        if current_section and is_multilevel and header_row and header_row_2 is None and row[0] in ['Project', 'TOTAL']:
            if row[0] == 'Project':
                header_row_2 = row
            continue
        
        # Capture data rows
        if current_section and header_row and row[0] not in [None, '', 'DEVELOPMENT', 'MAINTENANCE']:
            if current_section == 'dev':
                dev_data.append(row)
            elif current_section == 'maint':
                maint_data.append(row)
    
    return dev_data, maint_data, header_row, header_row_2, metadata


def _create_multilevel_columns():
    """Create multi-level column headers for weekly breakdown"""
    month_names_short = [month_name[i][:3] for i in range(1, 13)]
    
    multi_columns = []
    multi_columns.append(('', 'Project'))
    multi_columns.append(('', 'Component'))
    
    # 12 months × 5 weeks = 60 week columns
    for month_abbr in month_names_short:
        for week in range(1, 6):
            multi_columns.append((month_abbr, f'W{week}'))
    
    multi_columns.append(('', 'Total'))
    return multi_columns


def _create_single_level_columns(header_row):
    """Create single-level column headers with unique names"""
    unique_headers = []
    seen = {}
    for i, col in enumerate(header_row):
        if col is None or col == '':
            col = f'Column_{i}'
        
        col_str = str(col)
        if col_str in seen:
            seen[col_str] += 1
            unique_headers.append(f"{col_str}_{seen[col_str]}")
        else:
            seen[col_str] = 0
            unique_headers.append(col_str)
    
    return unique_headers


def _display_metadata_info(metadata: dict):
    """Display metadata information in a consistent format"""
    if not metadata or 'generated' not in metadata:
        return
    
    # Display timestamp (already includes Malaysia Time label from export)
    st.info(f":clock3: {metadata['generated']}")
    st.markdown("---")


def display_monthly_breakdown_preview(xlsx_path: Path, report_type: str = "monthly"):
    """Display monthly/weekly breakdown preview with team member selector"""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(xlsx_path, read_only=False)
        sheet_names = wb.sheetnames
        
        if not sheet_names:
            st.warning("No data found in the report")
            return
        
        # Team member selector
        breakdown_type = "weekly" if report_type == "weekly" else "monthly"
        selected_member = st.selectbox(
            "Select Team Member",
            options=sheet_names,
            help=f"Choose a team member to view their {breakdown_type} breakdown"
        )
        
        # Read the selected sheet
        ws = wb[selected_member]
        is_multilevel = report_type == "weekly"
        
        # Parse sections with metadata
        dev_data, maint_data, header_row, header_row_2, metadata = _parse_xlsx_sheet(ws, is_multilevel)
        
        # Display metadata
        _display_metadata_info(metadata)
        
        # Display Development table
        if dev_data and header_row:
            st.markdown("### :wrench: Development")
            
            if is_multilevel and header_row_2:
                multi_columns = _create_multilevel_columns()
                dev_df = pd.DataFrame(dev_data)
                dev_df.columns = pd.MultiIndex.from_tuples(multi_columns)
                project_col = ('', 'Project')
            else:
                unique_headers = _create_single_level_columns(header_row)
                dev_df = pd.DataFrame(dev_data, columns=unique_headers)
                project_col = next((c for c in dev_df.columns if str(c).startswith('Project')), 'Project')
            
            _display_dataframe_with_styling(dev_df, project_col, is_multilevel and header_row_2)
        
        # Display Maintenance table
        if maint_data and header_row:
            st.markdown("### :hammer_and_wrench: Maintenance")
            
            if is_multilevel and header_row_2:
                multi_columns = _create_multilevel_columns()
                maint_df = pd.DataFrame(maint_data)
                maint_df.columns = pd.MultiIndex.from_tuples(multi_columns)
                project_col = ('', 'Project')
            else:
                unique_headers = _create_single_level_columns(header_row)
                maint_df = pd.DataFrame(maint_data, columns=unique_headers)
                project_col = next((c for c in maint_df.columns if str(c).startswith('Project')), 'Project')
            
            _display_dataframe_with_styling(maint_df, project_col, is_multilevel and header_row_2)
        
        wb.close()
        
    except Exception as e:
        st.warning(f"Could not display monthly breakdown preview: {e}")
        logger.exception("Monthly breakdown preview failed")


def display_report_preview(result_path: Path, csv_data: bytes, report_type: str = "yearly", xlsx_path: Path = None):
    """Display report preview with separate tables for Development and Maintenance"""
    # Display title with report type
    report_type_display = {
        "yearly": "Yearly Overview",
        "quarterly": "Quarterly Breakdown",
        "monthly": "Monthly Breakdown",
        "weekly": "Weekly Breakdown"
    }.get(report_type, "Report")
    
    st.subheader(f":clipboard: Report Preview - {report_type_display}")
    
    # For monthly or weekly breakdown, show XLSX preview with team member selector
    if report_type in ["monthly", "weekly"] and xlsx_path and Path(xlsx_path).exists():
        display_monthly_breakdown_preview(xlsx_path, report_type)
        return
    
    try:
        # Parse the split CSV
        dev_df, maint_df, metadata = parse_split_csv(result_path)
        
        # Display metadata using shared function
        _display_metadata_info(metadata)
        stats = calculate_summary_stats(dev_df, maint_df)
        
        # Display summary metrics in two rows
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Projects", stats['projects'])
        with col2:
            st.metric("Components", stats['components'])
        with col3:
            st.metric("Team Members", stats['team_members'])
        
        # Display hours breakdown in one row
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Development Hours", f"{stats['dev_hours']:.1f}h")
        with col2:
            st.metric("Maintenance Hours", f"{stats['maint_hours']:.1f}h")
        with col3:
            st.metric("Total Hours", f"{stats['total_hours']:.1f}h")
        
        # Display Development table
        if not dev_df.empty:
            st.markdown("### :wrench: Development")
            dev_display = transform_to_multiindex(dev_df)
            
            def highlight_total(row):
                if isinstance(dev_display.columns, pd.MultiIndex):
                    is_total = row[('', 'Project')] == 'TOTAL'
                else:
                    is_total = row['Project'] == 'TOTAL'
                return ['background-color: #f0f2f6; font-weight: bold' if is_total else '' for _ in row]
            
            styled_dev = dev_display.style.apply(highlight_total, axis=1)
            
            # Format numeric columns
            if isinstance(dev_display.columns, pd.MultiIndex):
                format_dict = {col: '{:.1f}' for col in dev_display.columns if col not in [('', 'Project'), ('', 'Component')]}
            else:
                format_dict = {col: '{:.1f}' for col in dev_display.columns if col not in ['Project', 'Component']}
            
            styled_dev = styled_dev.format(format_dict, na_rep='-')
            
            # Create column config to pin Project and Component columns
            column_config = {}
            if isinstance(dev_display.columns, pd.MultiIndex):
                # For MultiIndex columns, use integer indices
                # Must account for DataFrame index levels even with hide_index=True
                num_index_levels = dev_display.index.nlevels
                column_config[num_index_levels] = st.column_config.Column(pinned=True)  # Project
                column_config[num_index_levels + 1] = st.column_config.Column(pinned=True)  # Component
            else:
                column_config['Project'] = st.column_config.Column(pinned=True)
                column_config['Component'] = st.column_config.Column(pinned=True)
                        
            st.dataframe(styled_dev, use_container_width=True, hide_index=True, height=300, column_config=column_config if column_config else None)
        
        # Display Maintenance table
        if not maint_df.empty:
            st.markdown("### :hammer_and_wrench: Maintenance")
            maint_display = transform_to_multiindex(maint_df)
            
            def highlight_total(row):
                if isinstance(maint_display.columns, pd.MultiIndex):
                    is_total = row[('', 'Project')] == 'TOTAL'
                else:
                    is_total = row['Project'] == 'TOTAL'
                return ['background-color: #f0f2f6; font-weight: bold' if is_total else '' for _ in row]
            
            styled_maint = maint_display.style.apply(highlight_total, axis=1)
            
            # Format numeric columns
            if isinstance(maint_display.columns, pd.MultiIndex):
                format_dict = {col: '{:.1f}' for col in maint_display.columns if col not in [('', 'Project'), ('', 'Component')]}
            else:
                format_dict = {col: '{:.1f}' for col in maint_display.columns if col not in ['Project', 'Component']}
            
            styled_maint = styled_maint.format(format_dict, na_rep='-')
            
            # Create column config to pin Project and Component columns
            column_config = {}
            if isinstance(maint_display.columns, pd.MultiIndex):
                # For MultiIndex columns, use integer indices
                # Must account for DataFrame index levels even with hide_index=True
                num_index_levels = maint_display.index.nlevels
                column_config[num_index_levels] = st.column_config.Column(pinned=True)  # Project
                column_config[num_index_levels + 1] = st.column_config.Column(pinned=True)  # Component
            else:
                column_config['Project'] = st.column_config.Column(pinned=True)
                column_config['Component'] = st.column_config.Column(pinned=True)
                        
            st.dataframe(styled_maint, use_container_width=True, hide_index=True, height=300, column_config=column_config if column_config else None)
        
    except Exception as e:
        st.warning(f"Could not display preview: {e}")
        logger.exception("Preview display failed")
        with st.expander(":page_facing_up: Raw CSV Preview"):
            st.code(csv_data.decode('utf-8')[:1000] + "\n...", language="csv")
