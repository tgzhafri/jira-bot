"""
Weekly breakdown exporter (CSV and XLSX formats) - one sheet per team member
"""

import csv
import logging
from pathlib import Path
from collections import defaultdict
from typing import Dict, List
from calendar import month_name

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

from .base_exporter import BaseExporter
from ..models import TimeEntry, ProjectComponent, WorkType, YearlyReport

logger = logging.getLogger(__name__)


class WeeklyBreakdownExporter(BaseExporter):
    """Export weekly breakdown reports - one sheet per team member"""
    
    def __init__(self, output_path: Path, filter_active_only: bool = True):
        """Initialize exporter with option to filter active employees only"""
        super().__init__(output_path)
        self.filter_active_only = filter_active_only
    
    def export_yearly(self, report: YearlyReport) -> Path:
        """Export yearly report with weekly breakdown per team member"""
        
        self._ensure_directory()
        
        # Aggregate data by author, month, and week
        # Structure: {Author: {WorkType: {ProjectComponent: {(month, week): hours}}}}
        from ..models import WorkType
        
        data_by_author = defaultdict(lambda: {
            WorkType.DEVELOPMENT: defaultdict(lambda: defaultdict(float)),
            WorkType.MAINTENANCE: defaultdict(lambda: defaultdict(float))
        })
        
        # Process entries to get week-level data
        for monthly_report in report.monthly_reports:
            month = monthly_report.month
            
            for entry in monthly_report.entries:
                # Filter by active status if specified
                if self.filter_active_only and not entry.author.active:
                    continue
                
                pc = entry.project_component
                author = entry.author
                work_type = entry.work_type
                
                # Only track Development and Maintenance
                if work_type in [WorkType.DEVELOPMENT, WorkType.MAINTENANCE]:
                    # Use week_hours from TimeEntry
                    if entry.week_hours:
                        for week, hours in entry.week_hours.items():
                            data_by_author[author][work_type][pc][(month, week)] += hours
                    else:
                        # Fallback: put all hours in week 1 if no week data
                        data_by_author[author][work_type][pc][(month, 1)] += entry.hours
        
        # Sort authors by display name
        sorted_authors = sorted(data_by_author.keys(), key=lambda a: a.display_name)
        
        # Export to CSV (simple format with author column)
        self._export_to_csv(data_by_author, sorted_authors, report)
        
        # Export to XLSX (one sheet per author with multi-level headers)
        xlsx_path = self._export_to_xlsx(data_by_author, sorted_authors, report)
        
        logger.info(f"Weekly breakdown reports exported")
        logger.info(f"  CSV: {self.output_path}")
        if xlsx_path:
            logger.info(f"  XLSX: {xlsx_path}")
        logger.info(f"  Team members: {len(sorted_authors)}")
        
        return (self.output_path, xlsx_path)
    
    def _export_to_csv(self, data_by_author: dict, sorted_authors: list, report: YearlyReport = None):
        """Export to CSV format with flat week headers (JanW1, JanW2, etc.)"""
        
        month_names = [month_name[i][:3] for i in range(1, 13)]
        
        # Build week headers: JanW1, JanW2, ..., DecW5
        week_headers = []
        for month_idx in range(1, 13):
            month_abbr = month_names[month_idx - 1]
            for week in range(1, 6):
                week_headers.append(f"{month_abbr}W{week}")
        
        with open(self.output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Write metadata header
            if report:
                self._write_metadata_header(writer, report)
            
            # Header
            header = ['Team Member', 'Work Type', 'Project', 'Component'] + week_headers + ['Total']
            writer.writerow(header)
            
            for author in sorted_authors:
                author_data = data_by_author[author]
                
                for work_type_name, work_type in [('Development', WorkType.DEVELOPMENT), ('Maintenance', WorkType.MAINTENANCE)]:
                    data = author_data[work_type]
                    sorted_pcs = sorted(data.keys(), key=lambda pc: (pc.project, pc.component.name))
                    
                    for pc in sorted_pcs:
                        row = [author.display_name.title(), work_type_name, pc.project, pc.component.name]
                        row_total = 0
                        
                        # Add hours for each week of each month
                        for month in range(1, 13):
                            for week in range(1, 6):
                                hours = data[pc].get((month, week), 0)
                                row.append(f"{hours:.1f}" if hours > 0 else "")
                                row_total += hours
                        
                        row.append(f"{row_total:.1f}" if row_total > 0 else "")
                        writer.writerow(row)
    
    def _write_weekly_section(self, ws, data: dict, sorted_pcs: list, row_start: int, section_name: str):
        """Write weekly data section with multi-level headers (reusable)"""
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        subheader_font = Font(bold=True, size=9)
        total_fill = PatternFill(start_color="F0F2F6", end_color="F0F2F6", fill_type="solid")
        total_font = Font(bold=True)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center', vertical='center')
        
        month_names = [month_name[i][:3] for i in range(1, 13)]
        row_idx = row_start
        
        # Section header
        ws.merge_cells(f'A{row_idx}:B{row_idx}')
        ws[f'A{row_idx}'] = section_name
        ws[f'A{row_idx}'].fill = section_fill
        ws[f'A{row_idx}'].font = section_font
        ws[f'A{row_idx}'].border = border
        row_idx += 1
        
        # Row 1: Month names (merged across 5 weeks each)
        ws[f'A{row_idx}'] = 'Project'
        ws[f'B{row_idx}'] = 'Component'
        ws[f'A{row_idx}'].fill = header_fill
        ws[f'A{row_idx}'].font = header_font
        ws[f'A{row_idx}'].border = border
        ws[f'B{row_idx}'].fill = header_fill
        ws[f'B{row_idx}'].font = header_font
        ws[f'B{row_idx}'].border = border
        
        col_idx = 3
        for month_idx, month_abbr in enumerate(month_names, start=1):
            # Merge cells for month name across 5 weeks
            start_col = get_column_letter(col_idx)
            end_col = get_column_letter(col_idx + 4)
            ws.merge_cells(f'{start_col}{row_idx}:{end_col}{row_idx}')
            ws[f'{start_col}{row_idx}'] = month_abbr
            ws[f'{start_col}{row_idx}'].fill = header_fill
            ws[f'{start_col}{row_idx}'].font = header_font
            ws[f'{start_col}{row_idx}'].alignment = center_align
            ws[f'{start_col}{row_idx}'].border = border
            col_idx += 5
        
        # Total column header (row 1)
        total_col = get_column_letter(col_idx)
        ws[f'{total_col}{row_idx}'] = 'Total'
        ws[f'{total_col}{row_idx}'].fill = header_fill
        ws[f'{total_col}{row_idx}'].font = header_font
        ws[f'{total_col}{row_idx}'].alignment = center_align
        ws[f'{total_col}{row_idx}'].border = border
        row_idx += 1
        
        # Row 2: Week numbers (W1-W5 for each month)
        ws[f'A{row_idx}'] = 'Project'
        ws[f'B{row_idx}'] = 'Component'
        ws[f'A{row_idx}'].fill = subheader_fill
        ws[f'A{row_idx}'].font = subheader_font
        ws[f'A{row_idx}'].border = border
        ws[f'B{row_idx}'].fill = subheader_fill
        ws[f'B{row_idx}'].font = subheader_font
        ws[f'B{row_idx}'].border = border
        
        col_idx = 3
        for month_idx in range(1, 13):
            for week in range(1, 6):
                col_letter = get_column_letter(col_idx)
                ws[f'{col_letter}{row_idx}'] = f'W{week}'
                ws[f'{col_letter}{row_idx}'].fill = subheader_fill
                ws[f'{col_letter}{row_idx}'].font = subheader_font
                ws[f'{col_letter}{row_idx}'].alignment = center_align
                ws[f'{col_letter}{row_idx}'].border = border
                col_idx += 1
        
        # Total column header (row 2)
        ws[f'{total_col}{row_idx}'] = 'Total'
        ws[f'{total_col}{row_idx}'].fill = subheader_fill
        ws[f'{total_col}{row_idx}'].font = subheader_font
        ws[f'{total_col}{row_idx}'].alignment = center_align
        ws[f'{total_col}{row_idx}'].border = border
        row_idx += 1
        
        # Data rows
        for pc in sorted_pcs:
            ws[f'A{row_idx}'] = pc.project
            ws[f'B{row_idx}'] = pc.component.name
            ws[f'A{row_idx}'].border = border
            ws[f'B{row_idx}'].border = border
            
            row_total = 0
            col_idx = 3
            for month in range(1, 13):
                for week in range(1, 6):
                    hours = data[pc].get((month, week), 0)
                    col_letter = get_column_letter(col_idx)
                    if hours > 0:
                        ws[f'{col_letter}{row_idx}'] = round(hours, 1)
                    ws[f'{col_letter}{row_idx}'].border = border
                    ws[f'{col_letter}{row_idx}'].alignment = Alignment(horizontal='right')
                    row_total += hours
                    col_idx += 1
            
            # Total
            if row_total > 0:
                ws[f'{total_col}{row_idx}'] = round(row_total, 1)
            ws[f'{total_col}{row_idx}'].border = border
            ws[f'{total_col}{row_idx}'].alignment = Alignment(horizontal='right')
            row_idx += 1
        
        # TOTAL row
        ws[f'A{row_idx}'] = 'TOTAL'
        ws[f'B{row_idx}'] = ''
        ws[f'A{row_idx}'].fill = total_fill
        ws[f'A{row_idx}'].font = total_font
        ws[f'A{row_idx}'].border = border
        ws[f'B{row_idx}'].fill = total_fill
        ws[f'B{row_idx}'].border = border
        
        col_idx = 3
        for month in range(1, 13):
            for week in range(1, 6):
                total_hours = sum(data[pc].get((month, week), 0) for pc in sorted_pcs)
                col_letter = get_column_letter(col_idx)
                if total_hours > 0:
                    ws[f'{col_letter}{row_idx}'] = round(total_hours, 1)
                ws[f'{col_letter}{row_idx}'].fill = total_fill
                ws[f'{col_letter}{row_idx}'].font = total_font
                ws[f'{col_letter}{row_idx}'].border = border
                ws[f'{col_letter}{row_idx}'].alignment = Alignment(horizontal='right')
                col_idx += 1
        
        # Grand total
        grand_total = sum(sum(data[pc].values()) for pc in sorted_pcs)
        if grand_total > 0:
            ws[f'{total_col}{row_idx}'] = round(grand_total, 1)
        ws[f'{total_col}{row_idx}'].fill = total_fill
        ws[f'{total_col}{row_idx}'].font = total_font
        ws[f'{total_col}{row_idx}'].border = border
        ws[f'{total_col}{row_idx}'].alignment = Alignment(horizontal='right')
        
        return row_idx + 1
    
    def _export_to_xlsx(self, data_by_author: dict, sorted_authors: list, report: YearlyReport = None) -> Path:
        """Export to XLSX format with one sheet per team member and multi-level headers"""
        
        if not XLSX_AVAILABLE:
            logger.warning("openpyxl not available - skipping XLSX export")
            return None
        
        xlsx_path = self.output_path.with_suffix('.xlsx')
        wb = Workbook()
        wb.remove(wb.active)
        
        for author in sorted_authors:
            # Create sheet for this author
            sheet_name = author.display_name.title()[:31]  # Excel sheet name limit
            ws = wb.create_sheet(title=sheet_name)
            
            # Add metadata header to XLSX
            current_row = 1
            if report and report.fetch_timestamp:
                timestamp_str = report.fetch_timestamp.strftime('%Y-%m-%d %H:%M:%S')
                ws[f'A{current_row}'] = f"Generated: {timestamp_str} (Malaysia Time)"
                current_row += 1
                current_row += 1  # Empty row
            
            author_data = data_by_author[author]
            
            # Development section
            dev_data = author_data[WorkType.DEVELOPMENT]
            sorted_dev_pcs = sorted(dev_data.keys(), key=lambda pc: (pc.project, pc.component.name))
            if sorted_dev_pcs:
                current_row = self._write_weekly_section(ws, dev_data, sorted_dev_pcs, current_row, "DEVELOPMENT")
                current_row += 1
            
            # Maintenance section
            maint_data = author_data[WorkType.MAINTENANCE]
            sorted_maint_pcs = sorted(maint_data.keys(), key=lambda pc: (pc.project, pc.component.name))
            if sorted_maint_pcs:
                self._write_weekly_section(ws, maint_data, sorted_maint_pcs, current_row, "MAINTENANCE")
            
            # Freeze panes (freeze first 2 rows of headers and 2 columns)
            ws.freeze_panes = 'C4'
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 25
            for col_idx in range(3, 64):  # 60 weeks + Total
                ws.column_dimensions[get_column_letter(col_idx)].width = 7
        
        wb.save(xlsx_path)
        logger.info(f"XLSX report exported to {xlsx_path}")
        
        return xlsx_path
    
    def export_monthly(self, report) -> Path:
        """Export monthly report - not used for this exporter"""
        raise NotImplementedError("Weekly breakdown exporter only supports yearly reports")
