"""
Quarterly breakdown exporter (CSV and XLSX formats)
"""

import csv
import logging
from pathlib import Path
from collections import defaultdict

from .base_exporter import BaseExporter
from ..models import YearlyReport, WorkType

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    logger.warning("openpyxl not available - XLSX export will be disabled")


class QuarterlyBreakdownExporter(BaseExporter):
    """Export quarterly breakdown reports to CSV format"""
    
    def __init__(self, output_path: Path, filter_active_only: bool = True):
        """Initialize exporter with option to filter active employees only"""
        super().__init__(output_path)
        self.filter_active_only = filter_active_only
    
    def _get_quarter(self, month: int) -> int:
        """Get quarter number (1-4) from month (1-12)"""
        return (month - 1) // 3 + 1
    
    def export_yearly(self, report: YearlyReport) -> Path:
        """Export yearly report to CSV with quarterly breakdown by project-component and author"""
        
        self._ensure_directory()
        
        # Aggregate data by quarter and author
        # Structure: {WorkType: {ProjectComponent: {Author: {quarter: hours}}}}
        from ..models import WorkType
        
        data_by_type = {
            WorkType.DEVELOPMENT: defaultdict(lambda: defaultdict(lambda: defaultdict(float))),
            WorkType.MAINTENANCE: defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
        }
        all_authors = set()
        
        for monthly_report in report.monthly_reports:
            quarter = self._get_quarter(monthly_report.month)
            
            for entry in monthly_report.entries:
                # Filter by active status if specified
                if self.filter_active_only and not entry.author.active:
                    continue
                
                pc = entry.project_component
                author = entry.author
                work_type = entry.work_type
                
                # Only track Development and Maintenance
                if work_type in [WorkType.DEVELOPMENT, WorkType.MAINTENANCE]:
                    data_by_type[work_type][pc][author][quarter] += entry.hours
                    all_authors.add(author)
        
        # Sort authors by display name
        sorted_authors = sorted(all_authors, key=lambda a: a.display_name)
        
        # Write CSV with separate sections
        with open(self.output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Write metadata header
            self._write_metadata_header(writer, report)
            
            # Header row - format names in Title Case with quarters
            header = ['Project', 'Component']
            for author in sorted_authors:
                for quarter in range(1, 5):
                    header.append(f"{author.display_name.title()} Q{quarter}")
            
            # Write Development section
            writer.writerow(['DEVELOPMENT'])
            writer.writerow(header)
            
            dev_data = data_by_type[WorkType.DEVELOPMENT]
            sorted_dev_pcs = sorted(dev_data.keys(), key=lambda pc: (pc.project, pc.component.name))
            
            for pc in sorted_dev_pcs:
                row = [pc.project, pc.component.name]
                for author in sorted_authors:
                    for quarter in range(1, 5):
                        hours = dev_data[pc].get(author, {}).get(quarter, 0)
                        row.append(f"{hours:.1f}" if hours > 0 else "")
                writer.writerow(row)
            
            # Development total row
            dev_total_row = ['TOTAL', '']
            for author in sorted_authors:
                for quarter in range(1, 5):
                    total_hours = sum(dev_data[pc].get(author, {}).get(quarter, 0) for pc in sorted_dev_pcs)
                    dev_total_row.append(f"{total_hours:.1f}" if total_hours > 0 else "")
            writer.writerow(dev_total_row)
            
            # Empty row separator
            writer.writerow([])
            
            # Write Maintenance section
            writer.writerow(['MAINTENANCE'])
            writer.writerow(header)
            
            maint_data = data_by_type[WorkType.MAINTENANCE]
            sorted_maint_pcs = sorted(maint_data.keys(), key=lambda pc: (pc.project, pc.component.name))
            
            for pc in sorted_maint_pcs:
                row = [pc.project, pc.component.name]
                for author in sorted_authors:
                    for quarter in range(1, 5):
                        hours = maint_data[pc].get(author, {}).get(quarter, 0)
                        row.append(f"{hours:.1f}" if hours > 0 else "")
                writer.writerow(row)
            
            # Maintenance total row
            maint_total_row = ['TOTAL', '']
            for author in sorted_authors:
                for quarter in range(1, 5):
                    total_hours = sum(maint_data[pc].get(author, {}).get(quarter, 0) for pc in sorted_maint_pcs)
                    maint_total_row.append(f"{total_hours:.1f}" if total_hours > 0 else "")
            writer.writerow(maint_total_row)
        
        logger.info(f"Quarterly CSV report exported to {self.output_path}")
        logger.info(f"  Development rows: {len(sorted_dev_pcs)}, Maintenance rows: {len(sorted_maint_pcs)}")
        
        # Also export to XLSX format with multi-level headers
        xlsx_path = self._export_to_xlsx(report, data_by_type, sorted_authors)
        
        # Return both paths as a tuple (CSV, XLSX)
        return (self.output_path, xlsx_path)
    
    def export_monthly(self, report) -> Path:
        """Export monthly report - not used for quarterly breakdown"""
        raise NotImplementedError("Quarterly breakdown exporter only supports yearly reports")

    
    def _create_header_rows(self, ws, sorted_authors: list, row_start: int = 1):
        """Create multi-level header rows with styling (reusable)"""
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        subheader_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Row 1: Team member names (merged across quarters)
        ws[f'A{row_start}'] = 'Project'
        ws[f'B{row_start}'] = 'Component'
        
        col_idx = 3
        for author in sorted_authors:
            # Merge cells for team member name across 4 quarters
            start_col = get_column_letter(col_idx)
            end_col = get_column_letter(col_idx + 3)
            ws.merge_cells(f'{start_col}{row_start}:{end_col}{row_start}')
            ws[f'{start_col}{row_start}'] = author.display_name.title()
            ws[f'{start_col}{row_start}'].fill = header_fill
            ws[f'{start_col}{row_start}'].font = header_font
            ws[f'{start_col}{row_start}'].alignment = center_align
            ws[f'{start_col}{row_start}'].border = border
            col_idx += 4
        
        # Row 2: Quarter headers
        row_sub = row_start + 1
        ws[f'A{row_sub}'] = 'Project'
        ws[f'B{row_sub}'] = 'Component'
        ws[f'A{row_sub}'].fill = subheader_fill
        ws[f'A{row_sub}'].font = subheader_font
        ws[f'A{row_sub}'].border = border
        ws[f'B{row_sub}'].fill = subheader_fill
        ws[f'B{row_sub}'].font = subheader_font
        ws[f'B{row_sub}'].border = border
        
        col_idx = 3
        for author in sorted_authors:
            for quarter in range(1, 5):
                col_letter = get_column_letter(col_idx)
                ws[f'{col_letter}{row_sub}'] = f'Q{quarter}'
                ws[f'{col_letter}{row_sub}'].fill = subheader_fill
                ws[f'{col_letter}{row_sub}'].font = subheader_font
                ws[f'{col_letter}{row_sub}'].alignment = center_align
                ws[f'{col_letter}{row_sub}'].border = border
                col_idx += 1
    
    def _write_data_section(self, ws, data: dict, sorted_pcs: list, sorted_authors: list, row_start: int, section_name: str = None):
        """Write data section with optional section header (reusable)"""
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        total_fill = PatternFill(start_color="F0F2F6", end_color="F0F2F6", fill_type="solid")
        total_font = Font(bold=True)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=12)
        
        row_idx = row_start
        
        # Optional section header
        if section_name:
            ws.merge_cells(f'A{row_idx}:B{row_idx}')
            ws[f'A{row_idx}'] = section_name
            ws[f'A{row_idx}'].fill = section_fill
            ws[f'A{row_idx}'].font = section_font
            ws[f'A{row_idx}'].border = border
            row_idx += 1
        
        # Data rows
        for pc in sorted_pcs:
            ws[f'A{row_idx}'] = pc.project
            ws[f'B{row_idx}'] = pc.component.name
            ws[f'A{row_idx}'].border = border
            ws[f'B{row_idx}'].border = border
            
            col_idx = 3
            for author in sorted_authors:
                for quarter in range(1, 5):
                    hours = data[pc].get(author, {}).get(quarter, 0)
                    col_letter = get_column_letter(col_idx)
                    if hours > 0:
                        ws[f'{col_letter}{row_idx}'] = round(hours, 1)
                    ws[f'{col_letter}{row_idx}'].border = border
                    ws[f'{col_letter}{row_idx}'].alignment = Alignment(horizontal='right')
                    col_idx += 1
            
            row_idx += 1
        
        # TOTAL row
        ws[f'A{row_idx}'] = 'TOTAL'
        ws[f'B{row_idx}'] = ''
        ws[f'A{row_idx}'].fill = total_fill
        ws[f'A{row_idx}'].font = total_font
        ws[f'A{row_idx}'].border = border
        ws[f'B{row_idx}'].fill = total_fill
        ws[f'B{row_idx}'].border = border
        
        col_idx = 3
        for author in sorted_authors:
            for quarter in range(1, 5):
                total_hours = sum(data[pc].get(author, {}).get(quarter, 0) for pc in sorted_pcs)
                col_letter = get_column_letter(col_idx)
                if total_hours > 0:
                    ws[f'{col_letter}{row_idx}'] = round(total_hours, 1)
                ws[f'{col_letter}{row_idx}'].fill = total_fill
                ws[f'{col_letter}{row_idx}'].font = total_font
                ws[f'{col_letter}{row_idx}'].border = border
                ws[f'{col_letter}{row_idx}'].alignment = Alignment(horizontal='right')
                col_idx += 1
        
        return row_idx + 1  # Return next available row
    
    def _export_to_xlsx(self, report: YearlyReport, data_by_type: dict, sorted_authors: list) -> Path:
        """Export to XLSX format with multi-level headers and styling"""
        
        if not XLSX_AVAILABLE:
            logger.warning("openpyxl not available - skipping XLSX export")
            return None
        
        # Change extension to .xlsx
        xlsx_path = self.output_path.with_suffix('.xlsx')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Quarterly Report"
        
        # Add metadata header to XLSX
        current_row = 1
        if report and report.fetch_timestamp:
            timestamp_str = report.fetch_timestamp.strftime('%Y-%m-%d %H:%M:%S')
            ws[f'A{current_row}'] = f"Generated: {timestamp_str} (Malaysia Time)"
            current_row += 1
            current_row += 1  # Empty row
        
        # Create header rows once at the top
        header_start = current_row
        self._create_header_rows(ws, sorted_authors, row_start=header_start)
        
        current_row = header_start + 2  # Start after headers (2 header rows)
        
        # Development section
        dev_data = data_by_type[WorkType.DEVELOPMENT]
        sorted_dev_pcs = sorted(dev_data.keys(), key=lambda pc: (pc.project, pc.component.name))
        current_row = self._write_data_section(ws, dev_data, sorted_dev_pcs, sorted_authors, current_row, "DEVELOPMENT")
        
        # Empty row separator
        current_row += 1
        
        # Maintenance section
        maint_data = data_by_type[WorkType.MAINTENANCE]
        sorted_maint_pcs = sorted(maint_data.keys(), key=lambda pc: (pc.project, pc.component.name))
        self._write_data_section(ws, maint_data, sorted_maint_pcs, sorted_authors, current_row, "MAINTENANCE")
        
        # Freeze panes (freeze first 2 rows and 2 columns)
        ws.freeze_panes = 'C3'
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        for col_idx in range(3, 3 + len(sorted_authors) * 4):
            ws.column_dimensions[get_column_letter(col_idx)].width = 10
        
        # Save workbook
        wb.save(xlsx_path)
        logger.info(f"XLSX report exported to {xlsx_path}")
        
        return xlsx_path
