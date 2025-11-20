#!/usr/bin/env python3
"""
Streamlit web UI for Automate Jira
"""

import streamlit as st
import sys
import logging
import pandas as pd
from pathlib import Path
from datetime import datetime

# Add src to path
sys.path.insert(0, str(Path(__file__).parent))

from src.config import Config
from src.jira_client import JiraClient
from src.utils import get_month_range, format_date_for_jql
from src.report_generator import generate_csv_report, generate_quarterly_report, generate_monthly_breakdown_report

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ============================================================================
# UI Components
# ============================================================================

def show_config_error(error_msg: str):
    """Display configuration error with helpful instructions"""
    st.error(":warning: Missing Configuration")
    st.markdown("""
    Please set these environment variables:
    ```
    JIRA_URL=https://your-company.atlassian.net
    JIRA_USERNAME=your-email@company.com
    JIRA_API_TOKEN=your-api-token
    ```
    
    **For Docker:** Make sure your `.env` file exists and restart:
    ```bash
    make web-stop
    make web
    ```
    """)
    st.error(f"Error details: {error_msg}")
    st.stop()


def parse_split_csv(file_path: Path) -> tuple:
    """Parse CSV file split by Development and Maintenance sections"""
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    dev_lines = []
    maint_lines = []
    current_section = None
    
    for line in lines:
        stripped = line.strip()
        if stripped == 'DEVELOPMENT':
            current_section = 'dev'
            continue
        elif stripped == 'MAINTENANCE':
            current_section = 'maint'
            continue
        elif not stripped:
            continue
        
        if current_section == 'dev':
            dev_lines.append(line)
        elif current_section == 'maint':
            maint_lines.append(line)
    
    # Parse into DataFrames
    from io import StringIO
    dev_df = pd.read_csv(StringIO(''.join(dev_lines))) if dev_lines else pd.DataFrame()
    maint_df = pd.read_csv(StringIO(''.join(maint_lines))) if maint_lines else pd.DataFrame()
    
    return dev_df, maint_df


def calculate_summary_stats(dev_df: pd.DataFrame, maint_df: pd.DataFrame) -> dict:
    """Calculate summary statistics from both Development and Maintenance DataFrames"""
    
    def get_stats_from_df(df):
        if df.empty:
            return 0, 0, 0
        data_df = df[df['Project'].notna() & (df['Project'] != 'TOTAL')]
        projects = data_df['Project'].nunique() if 'Project' in data_df.columns else 0
        components = data_df['Component'].nunique() if 'Component' in data_df.columns else 0
        
        # Get total hours from TOTAL row
        total_row = df[df['Project'] == 'TOTAL']
        if not total_row.empty:
            numeric_cols = total_row.select_dtypes(include=['float64', 'int64']).columns
            hours = total_row[numeric_cols].sum().sum()
        else:
            numeric_cols = data_df.select_dtypes(include=['float64', 'int64']).columns
            hours = data_df[numeric_cols].sum().sum()
        
        return projects, components, hours
    
    dev_projects, dev_components, dev_hours = get_stats_from_df(dev_df)
    maint_projects, maint_components, maint_hours = get_stats_from_df(maint_df)
    
    # Team members count - for quarterly reports, divide by 4 (Q1-Q4 per member)
    team_members = 0
    if not dev_df.empty:
        # Check if this is a quarterly report (columns end with Q1, Q2, Q3, Q4)
        cols = [c for c in dev_df.columns if c not in ['Project', 'Component']]
        if cols and any('Q' in str(col) for col in cols):
            # Quarterly report: each member has 4 columns (Q1-Q4)
            team_members = len(cols) // 4
        else:
            # Yearly report: one column per member
            team_members = len(cols)
    elif not maint_df.empty:
        cols = [c for c in maint_df.columns if c not in ['Project', 'Component']]
        if cols and any('Q' in str(col) for col in cols):
            team_members = len(cols) // 4
        else:
            team_members = len(cols)
    
    return {
        'projects': max(dev_projects, maint_projects),
        'components': dev_components + maint_components,
        'team_members': team_members,
        'dev_hours': dev_hours,
        'maint_hours': maint_hours,
        'total_hours': dev_hours + maint_hours
    }


def transform_to_multiindex(df: pd.DataFrame) -> pd.DataFrame:
    """Transform quarterly columns to multi-level index for better display"""
    if df.empty:
        return df
    
    # Check if this is a quarterly report
    cols = [c for c in df.columns if c not in ['Project', 'Component']]
    if not cols or not any('Q' in str(col) for col in cols):
        return df  # Not quarterly, return as-is
    
    # Create multi-level columns
    new_columns = [('', 'Project'), ('', 'Component')]
    
    for col in cols:
        # Parse "Name Q1" format
        if ' Q' in col:
            parts = col.rsplit(' Q', 1)
            name = parts[0]
            quarter = f'Q{parts[1]}'
            new_columns.append((name, quarter))
        else:
            new_columns.append(('', col))
    
    # Create new dataframe with multi-level columns
    df_multi = df.copy()
    df_multi.columns = pd.MultiIndex.from_tuples(new_columns)
    
    return df_multi


def display_monthly_breakdown_preview(xlsx_path: Path):
    """Display monthly breakdown preview with team member selector"""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(xlsx_path, read_only=True)
        sheet_names = wb.sheetnames
        
        if not sheet_names:
            st.warning("No data found in the report")
            return
        
        # Team member selector
        selected_member = st.selectbox(
            "Select Team Member",
            options=sheet_names,
            help="Choose a team member to view their monthly breakdown"
        )
        
        # Read the selected sheet
        ws = wb[selected_member]
        
        # Parse sections manually to avoid column name conflicts
        dev_data = []
        maint_data = []
        current_section = None
        header_row = None
        
        for row in ws.iter_rows(values_only=True):
            # Skip empty rows
            if not any(row):
                continue
            
            # Check for section headers
            if row[0] == 'DEVELOPMENT':
                current_section = 'dev'
                header_row = None
                continue
            elif row[0] == 'MAINTENANCE':
                current_section = 'maint'
                header_row = None
                continue
            
            # Capture header row for each section
            if current_section and header_row is None and row[0] in ['Project', 'TOTAL']:
                if row[0] == 'Project':
                    header_row = row
                continue
            
            # Capture data rows
            if current_section and header_row and row[0] not in [None, '', 'DEVELOPMENT', 'MAINTENANCE']:
                if current_section == 'dev':
                    dev_data.append(row)
                elif current_section == 'maint':
                    maint_data.append(row)
        
        # Display tables with TOTAL row styling
        if dev_data and header_row:
            st.markdown("### :wrench: Development")
            dev_df = pd.DataFrame(dev_data, columns=header_row)
            
            # Style TOTAL row
            def highlight_total(row):
                return ['background-color: #f0f2f6; font-weight: bold' if row['Project'] == 'TOTAL' else '' for _ in row]
            
            # Format numeric columns to 1 decimal place (custom formatter to handle None/empty)
            def format_number(val):
                if val is None or val == '' or val == '-':
                    return '-'
                try:
                    return f'{float(val):.1f}'
                except (ValueError, TypeError):
                    return str(val)
            
            format_dict = {col: format_number for col in dev_df.columns if col not in ['Project', 'Component']}
            
            styled_dev = dev_df.style.apply(highlight_total, axis=1).format(format_dict)
            st.dataframe(styled_dev, use_container_width=True, hide_index=True, height=250)
        
        if maint_data and header_row:
            st.markdown("### :hammer_and_wrench: Maintenance")
            maint_df = pd.DataFrame(maint_data, columns=header_row)
            
            # Style TOTAL row
            def highlight_total(row):
                return ['background-color: #f0f2f6; font-weight: bold' if row['Project'] == 'TOTAL' else '' for _ in row]
            
            # Format numeric columns to 1 decimal place (custom formatter to handle None/empty)
            def format_number(val):
                if val is None or val == '' or val == '-':
                    return '-'
                try:
                    return f'{float(val):.1f}'
                except (ValueError, TypeError):
                    return str(val)
            
            format_dict = {col: format_number for col in maint_df.columns if col not in ['Project', 'Component']}
            
            styled_maint = maint_df.style.apply(highlight_total, axis=1).format(format_dict)
            st.dataframe(styled_maint, use_container_width=True, hide_index=True, height=250)
        
        wb.close()
        
    except Exception as e:
        st.warning(f"Could not display monthly breakdown preview: {e}")
        logger.exception("Monthly breakdown preview failed")


def display_report_preview(result_path: Path, csv_data: bytes, report_type: str = "yearly", xlsx_path: Path = None):
    """Display report preview with separate tables for Development and Maintenance"""
    # Display title with report type
    report_type_display = {
        "yearly": "Yearly Overview",
        "quarterly": "Quarterly Breakdown",
        "monthly": "Monthly Breakdown"
    }.get(report_type, "Report")
    
    st.subheader(f":clipboard: Report Preview - {report_type_display}")
    
    # For monthly breakdown, show XLSX preview with team member selector
    if report_type == "monthly" and xlsx_path and Path(xlsx_path).exists():
        display_monthly_breakdown_preview(xlsx_path)
        return
    
    try:
        # Parse the split CSV
        dev_df, maint_df = parse_split_csv(result_path)
        stats = calculate_summary_stats(dev_df, maint_df)
        
        # Display summary metrics in two rows
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Projects", stats['projects'])
        with col2:
            st.metric("Components", stats['components'])
        with col3:
            st.metric("Team Members", stats['team_members'])
        
        # Display hours breakdown in one row
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Development Hours", f"{stats['dev_hours']:.1f}h")
        with col2:
            st.metric("Maintenance Hours", f"{stats['maint_hours']:.1f}h")
        with col3:
            st.metric("Total Hours", f"{stats['total_hours']:.1f}h")
        
        # Display Development table
        if not dev_df.empty:
            st.markdown("### :wrench: Development")
            # Transform to multi-level columns for quarterly reports
            dev_display = transform_to_multiindex(dev_df)
            
            # Style the TOTAL row and format numbers to 1 decimal place
            def highlight_total(row):
                # Check if Project column is TOTAL
                if isinstance(dev_display.columns, pd.MultiIndex):
                    is_total = row[('', 'Project')] == 'TOTAL'
                else:
                    is_total = row['Project'] == 'TOTAL'
                return ['background-color: #f0f2f6; font-weight: bold' if is_total else '' for _ in row]
            
            styled_dev = dev_display.style.apply(highlight_total, axis=1)
            
            # Format numeric columns
            if isinstance(dev_display.columns, pd.MultiIndex):
                format_dict = {col: '{:.1f}' for col in dev_display.columns if col not in [('', 'Project'), ('', 'Component')]}
            else:
                format_dict = {col: '{:.1f}' for col in dev_display.columns if col not in ['Project', 'Component']}
            
            styled_dev = styled_dev.format(format_dict, na_rep='-')
                        
            st.dataframe(
                styled_dev,
                use_container_width=True,
                hide_index=True,
                height=300
            )
        
        # Display Maintenance table
        if not maint_df.empty:
            st.markdown("### :hammer_and_wrench: Maintenance")
            # Transform to multi-level columns for quarterly reports
            maint_display = transform_to_multiindex(maint_df)
            
            # Style the TOTAL row and format numbers to 1 decimal place
            def highlight_total(row):
                if isinstance(maint_display.columns, pd.MultiIndex):
                    is_total = row[('', 'Project')] == 'TOTAL'
                else:
                    is_total = row['Project'] == 'TOTAL'
                return ['background-color: #f0f2f6; font-weight: bold' if is_total else '' for _ in row]
            
            styled_maint = maint_display.style.apply(highlight_total, axis=1)
            
            # Format numeric columns
            if isinstance(maint_display.columns, pd.MultiIndex):
                format_dict = {col: '{:.1f}' for col in maint_display.columns if col not in [('', 'Project'), ('', 'Component')]}
            else:
                format_dict = {col: '{:.1f}' for col in maint_display.columns if col not in ['Project', 'Component']}
            
            styled_maint = styled_maint.format(format_dict, na_rep='-')
                        
            st.dataframe(
                styled_maint,
                use_container_width=True,
                hide_index=True,
                height=300
            )
        
    except Exception as e:
        st.warning(f"Could not display preview: {e}")
        logger.exception("Preview display failed")
        with st.expander(":page_facing_up: Raw CSV Preview"):
            st.code(csv_data.decode('utf-8')[:1000] + "\n...", language="csv")





# ============================================================================
# Main Application
# ============================================================================

def main():
    """Main application entry point"""
    
    try:
        # Page config
        st.set_page_config(
            page_title="Automate Jira",
            page_icon=":bar_chart:",
            layout="wide"
        )
        
        # Header
        st.title(":bar_chart: Automate Jira")
        st.markdown("Generate CSV reports of team hours by project and component")
    except Exception as e:
        st.error(f":x: Error initializing page: {e}")
        logger.exception("Page initialization failed")
        return
    
    # Initialize session state for report persistence
    if 'report_generated' not in st.session_state:
        st.session_state.report_generated = False
    if 'csv_path' not in st.session_state:
        st.session_state.csv_path = None
    if 'xlsx_path' not in st.session_state:
        st.session_state.xlsx_path = None
    if 'report_type' not in st.session_state:
        st.session_state.report_type = None
    if 'csv_data' not in st.session_state:
        st.session_state.csv_data = None
    
    # Sidebar configuration
    st.sidebar.header(":gear: Configuration")
    
    # Report type selector
    report_type = st.sidebar.radio(
        "Report Type",
        options=["Yearly Overview", "Quarterly Breakdown", "Monthly Breakdown"],
        index=0,
        help="Choose report type: yearly summary, quarterly breakdown, or monthly breakdown per team member"
    )
    
    current_year = datetime.now().year
    year = st.sidebar.number_input(
        "Report Year",
        min_value=2020,
        max_value=current_year + 1,
        value=current_year,
        step=1
    )
    
    with st.sidebar.expander("Advanced Options"):
        max_workers = st.slider(
            "Parallel Workers",
            min_value=1,
            max_value=16,
            value=8,
            help="Number of parallel threads for fetching data"
        )
        
        use_cache = st.checkbox(
            "Enable Cache",
            value=True,
            help="Cache API responses for faster subsequent runs"
        )
        
        # Cache management - will be handled after config is loaded
        clear_cache_clicked = st.button(":wastebasket: Clear Cache", help="Delete all cached API responses")
    
    st.markdown("---")
    
    # Load and validate configuration
    try:
        config = Config.from_env()
        config.jira.enable_cache = use_cache
        config.jira.max_workers = max_workers
        
        # Ensure cache directory is absolute path
        if not Path(config.jira.cache_dir).is_absolute():
            config.jira.cache_dir = str(Path.cwd() / config.jira.cache_dir)
        
        config.validate()
        
        # Show connection info in compact format
        cache_path = Path(config.jira.cache_dir)
        cache_files = list(cache_path.glob("*.json")) if cache_path.exists() else []
        
        if config.jira.enable_cache and cache_files:
            cache_status = f"{len(cache_files)} cached"
        elif not config.jira.enable_cache:
            cache_status = "disabled"
        else:
            cache_status = "empty"
        
        # Display connection info
        st.info(f":link: Connected: {config.jira.url} | :floppy_disk: Cache: {cache_status}")
        
        # Handle cache clearing after config is loaded
        if clear_cache_clicked:
            if cache_path.exists() and cache_path.is_dir():
                import shutil
                # Clear contents without removing directory
                for item in cache_path.iterdir():
                    if item.is_file():
                        item.unlink()
                    elif item.is_dir():
                        shutil.rmtree(item)
                st.success(":white_check_mark: Cache cleared!")
                st.rerun()
            else:
                st.info("Cache directory does not exist")
        
    except ValueError as e:
        show_config_error(str(e))
        return
    except Exception as e:
        st.error(f":x: Configuration error: {e}")
        st.stop()
        return
    
    # Generate button
    button_label = ":rocket: Generate Report"
    if st.button(button_label, type="primary", use_container_width=True):
        # Determine report type and settings
        is_quarterly = report_type == "Quarterly Breakdown"
        is_monthly = report_type == "Monthly Breakdown"
        
        if is_monthly:
            report_name = "monthly breakdown"
        elif is_quarterly:
            report_name = "quarterly breakdown"
        else:
            report_name = "yearly overview"
        
        with st.spinner(f"Generating {report_name} report for {year}..."):
            try:
                progress_text = st.empty()
                progress_bar = st.progress(0)
                
                Path("reports").mkdir(exist_ok=True)
                
                progress_text.text("Fetching data from Jira...")
                progress_bar.progress(30)
                
                # Generate report based on type
                if is_monthly:
                    output_file = f"reports/monthly_breakdown_{year}.csv"
                    result = generate_monthly_breakdown_report(
                        config,
                        year=year,
                        output_file=output_file,
                        max_workers=max_workers
                    )
                    csv_path, xlsx_path = result
                    result_path = csv_path
                elif is_quarterly:
                    output_file = f"reports/quarterly_report_{year}.csv"
                    result = generate_quarterly_report(
                        config,
                        year=year,
                        output_file=output_file,
                        max_workers=max_workers
                    )
                    csv_path, xlsx_path = result
                    result_path = csv_path
                else:
                    output_file = f"reports/manhour_report_{year}.csv"
                    result_path = generate_csv_report(
                        config,
                        year=year,
                        output_file=output_file,
                        max_workers=max_workers
                    )
                    xlsx_path = None
                
                # Clear progress indicators
                progress_bar.empty()
                progress_text.empty()
                
                if result_path and Path(result_path).exists():
                    # Store in session state
                    st.session_state.report_generated = True
                    st.session_state.csv_path = result_path
                    st.session_state.xlsx_path = xlsx_path
                    st.session_state.report_type = report_type
                    
                    # Read CSV data
                    with open(result_path, 'rb') as f:
                        st.session_state.csv_data = f.read()
                else:
                    st.warning(":warning: No data found for the specified period")
                    
            except Exception as e:
                st.error(f":x: Error generating report: {e}")
                logger.exception("Report generation failed")
    
    # Display report if it exists in session state AND matches current report type selection
    if st.session_state.report_generated and st.session_state.csv_path:
        csv_path = st.session_state.csv_path
        xlsx_path = st.session_state.xlsx_path
        csv_data = st.session_state.csv_data
        report_type_stored = st.session_state.report_type
        
        # Only display if the stored report type matches the current selection
        if report_type_stored == report_type and Path(csv_path).exists():
            # Show download buttons
            is_quarterly = report_type_stored == "Quarterly Breakdown"
            is_monthly = report_type_stored == "Monthly Breakdown"
            
            if (is_quarterly or is_monthly) and xlsx_path and Path(xlsx_path).exists():
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        label=":inbox_tray: Download CSV",
                        data=csv_data,
                        file_name=Path(csv_path).name,
                        mime="text/csv",
                        use_container_width=True,
                        key="download_csv"
                    )
                with col2:
                    with open(xlsx_path, 'rb') as f:
                        xlsx_data = f.read()
                    st.download_button(
                        label=":inbox_tray: Download XLSX (Formatted)",
                        data=xlsx_data,
                        file_name=Path(xlsx_path).name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="download_xlsx"
                    )
            else:
                st.download_button(
                    label=":inbox_tray: Download CSV Report",
                    data=csv_data,
                    file_name=Path(csv_path).name,
                    mime="text/csv",
                    use_container_width=True,
                    key="download_csv"
                )
            
            # Display preview based on report type
            if is_monthly:
                display_report_preview(Path(csv_path), csv_data, report_type="monthly", xlsx_path=Path(xlsx_path) if xlsx_path else None)
            elif is_quarterly:
                display_report_preview(Path(csv_path), csv_data, report_type="quarterly", xlsx_path=Path(xlsx_path) if xlsx_path else None)
            else:
                display_report_preview(Path(csv_path), csv_data, report_type="yearly")
    
    # Footer
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: #666; font-size: 0.9em;'>
        Built using Streamlit | Automate Jira v1.0.0
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
